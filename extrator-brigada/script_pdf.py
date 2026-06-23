import firebase_admin
from firebase_admin import credentials, firestore
from fpdf import FPDF
from datetime import datetime
import pandas as pd
import os
import json
import re

# 1. Configuração do Firebase
service_account_path = "serviceAccountKey.json"
if not os.path.exists(service_account_path):
    service_account_path = "extrator-brigada/serviceAccountKey.json"

if os.path.exists(service_account_path):
    cred = credentials.Certificate(service_account_path)
elif "FIREBASE_SERVICE_ACCOUNT_JSON" in os.environ:
    try:
        info = json.loads(os.environ["FIREBASE_SERVICE_ACCOUNT_JSON"])
        cred = credentials.Certificate(info)
    except Exception as e:
        print(f"Erro ao carregar credenciais da variável FIREBASE_SERVICE_ACCOUNT_JSON: {e}")
        cred = None
else:
    try:
        cred = credentials.ApplicationDefault()
    except Exception as e:
        print(f"Aviso: Não foi possível obter credenciais padrão ({e}).")
        cred = None

if cred:
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred)
    db = firestore.client()
else:
    raise Exception("Erro crítico: Nenhuma credencial do Firebase configurada.")

# 2. Puxar dados do inventário e vistorias
print("Carregando inventário...")
inventario_ref = db.collection('artifacts', 'brigadantb', 'public', 'data', 'inventario')
inventario_docs = inventario_ref.stream()
map_inventario = {}
for doc in inventario_docs:
    inv_data = doc.to_dict()
    map_inventario[doc.id.strip().upper()] = {
        'local': inv_data.get('local', 'Local não especificado'),
        'tipo': inv_data.get('tipoKgL') or inv_data.get('tipo') or 'Tipo não especificado',
        'categoria': inv_data.get('categoria', 'Extintor')
    }

print("Carregando inspeções...")
inspecoes_ref = db.collection('artifacts', 'brigadantb', 'public', 'data', 'inspecoes')
docs = inspecoes_ref.stream()

lista_extintores = []
lista_hidrantes = []

for doc in docs:
    data = doc.to_dict()
    id_ext = data.get('idExtintor', '')
    if not id_ext or 'teste' in id_ext.lower():
        continue
    
    ext_key = id_ext.strip().upper()
    inv_info = map_inventario.get(ext_key, {'local': 'Local não cadastrado', 'tipo': 'Tipo não cadastrado', 'categoria': 'Extintor'})
    data['local'] = inv_info['local']
    data['tipo'] = inv_info['tipo']
    categoria = inv_info['categoria']
    
    if categoria == 'Hidrante':
        lista_hidrantes.append(data)
    else:
        lista_extintores.append(data)

def chave_ordenacao(item):
    id_ext = item.get('idExtintor', '')
    match = re.search(r'\d+', id_ext)
    return int(match.group()) if match else 9999

lista_extintores.sort(key=chave_ordenacao)
lista_hidrantes.sort(key=chave_ordenacao)

MAP_CHECKLIST_EXT = {
    'acesso': 'Desobstrucao e Acesso',
    'sinalizacaoParede': 'Sinalizacao de Parede (Placa)',
    'sinalizacaoPiso': 'Sinalizacao de Piso (Pintura)',
    'suporte': 'Suporte e Altura de Fixacao',
    'cilindro': 'Casco / Cilindro Sem Corrosao',
    'instrucoes': 'Quadro de Instrucoes Legivel',
    'mangueira': 'Mangueira e Bico / Difusor',
    'lacre': 'Lacre de Seguranca Intacto',
    'trava': 'Trava de Seguranca / Pino',
    'manometro': 'Manometro na Faixa Verde'
}

MAP_CHECKLIST_HID = {
    'caixa_abrigo': 'Caixa do Abrigo',
    'sinal_piso': 'Sinalizacao de Piso 1m X 1m',
    'sinal_placa': 'Sinalizacao Placa 1,80cm',
    'qtd_mangueira': 'Quantidade de Mangueira',
    'esguicho': 'Esguicho Regulavel',
    'valvula_globo': 'Valvula de Globo',
    'chave_storz': 'Chave Storz',
    'adaptador': 'Adaptador de 2 1/2 P/ 1 1/2',
    'acrilico': 'Acrilico',
    'prox_teste_hidro': 'Prox. Teste Hidrostatico'
}

class RelatorioPDF(FPDF):
    def header(self):
        self.set_fill_color(22, 30, 49)
        self.rect(0, 0, 210, 32, 'F')
        self.set_fill_color(239, 68, 68)
        self.rect(0, 31, 210, 1, 'F')
        self.set_xy(10, 8)
        self.set_text_color(255, 255, 255)
        self.set_font('Helvetica', 'B', 15)
        self.cell(0, 6, 'NIBT BRIGADA - RELATORIO DE VISTORIAS', 0, 1, 'L')
        self.set_font('Helvetica', '', 9)
        self.set_text_color(203, 213, 225)
        self.cell(0, 4, 'Controle de Inventario, Validade e Conformidade', 0, 1, 'L')
        self.set_xy(150, 12)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(148, 163, 184)
        self.cell(50, 4, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 0, 'R')
        self.set_y(38)

    def footer(self):
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(100, 116, 139)
        self.cell(0, 10, 'NIBT Brigada - Painel Operacional', 0, 0, 'L')
        self.cell(0, 10, f'Pagina {self.page_no()}/{{nb}}', 0, 0, 'R')

def gerar_secao_pdf(pdf, titulo, lista_dados, map_checklist):
    conformes = []
    nao_conformes = []
    for item in lista_dados:
        conformidade = item.get('conformidade', {})
        if not conformidade or any(v == 'Não Conforme' for v in conformidade.values()):
            nao_conformes.append(item)
        else:
            conformes.append(item)
            
    total = len(lista_dados)
    qtd_conf = len(conformes)
    qtd_nconf = len(nao_conformes)
    taxa = f"{int(qtd_conf / total * 100)}%" if total > 0 else "0%"

    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 10, titulo.upper(), 0, 1, 'L')
    pdf.ln(2)

    y_start = pdf.get_y()
    pdf.set_fill_color(248, 250, 252)
    pdf.set_draw_color(226, 232, 240)
    pdf.rect(10, y_start, 190, 22, 'DF')

    pdf.set_xy(10, y_start + 2)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(47, 4, 'TOTAL INSPECIONADOS', 0, 0, 'C')
    pdf.cell(47, 4, 'EM CONFORMIDADE', 0, 0, 'C')
    pdf.cell(47, 4, 'NAO CONFORMES', 0, 0, 'C')
    pdf.cell(49, 4, 'TAXA DE CONFORMIDADE', 0, 1, 'C')

    pdf.set_x(10)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(47, 8, str(total), 0, 0, 'C')
    pdf.set_text_color(21, 128, 61)
    pdf.cell(47, 8, str(qtd_conf), 0, 0, 'C')
    pdf.set_text_color(185, 28, 28)
    pdf.cell(47, 8, str(qtd_nconf), 0, 0, 'C')
    pdf.set_text_color(30, 41, 59)
    pdf.cell(49, 8, taxa, 0, 1, 'C')
    
    pdf.set_y(y_start + 30)

    # NÃO CONFORMES
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_text_color(185, 28, 28)
    pdf.cell(0, 6, f'1. DETALHES DE {titulo.upper()} NAO CONFORMES', 0, 1, 'L')
    pdf.ln(2)

    if not nao_conformes:
        pdf.set_font('Helvetica', 'I', 10)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 8, f'Nenhum {titulo.lower()} apresenta pendencias.', 0, 1, 'L')
        pdf.ln(5)
    else:
        for item in nao_conformes:
            altura_estimada = 20
            conformidade = item.get('conformidade', {})
            pendencias = [k for k, v in conformidade.items() if v == 'Não Conforme']
            if pendencias:
                altura_estimada += 5 + (4 * len(pendencias))
            if item.get('observacoes', '').strip():
                altura_estimada += 5
                
            if pdf.get_y() + altura_estimada > 275:
                pdf.add_page()

            pdf.set_fill_color(254, 226, 226)
            pdf.set_draw_color(252, 165, 165)
            pdf.set_text_color(153, 27, 27)
            pdf.set_font('Helvetica', 'B', 9)
            
            id_el = item.get('idExtintor', '')
            local = item.get('local', 'N/A')
            tipo = item.get('tipo', 'N/A')
            brigadista = item.get('nomeBrigadista', 'N/A')
            data_insp = item.get('dataInspecao', 'N/A')
            try:
                data_insp = datetime.strptime(data_insp, '%Y-%m-%d').strftime('%d/%m/%Y')
            except:
                pass
                
            pdf.cell(190, 7, f" {id_el} - {local}  ({tipo})  |  Brigadista: {brigadista}  |  Data: {data_insp}", 1, 1, 'L', True)
            
            pdf.set_text_color(51, 65, 85)
            pdf.set_font('Helvetica', '', 8.5)
            venc_recarga = item.get('vencimentoRecarga', 'N/A')
            venc_hidro = item.get('vencimentoHidrostatico', 'N/A')
            try: venc_recarga = datetime.strptime(venc_recarga, '%Y-%m-%d').strftime('%d/%m/%Y')
            except: pass
            try: venc_hidro = datetime.strptime(venc_hidro, '%Y-%m-%d').strftime('%d/%m/%Y')
            except: pass
                
            pdf.cell(190, 6, f"  Vencimento Recarga/Manutencao: {venc_recarga}   |   Teste Hidrostatico: {venc_hidro}", 'LR', 1, 'L')
            
            pendencias_mapped = [map_checklist.get(k, k) for k, v in conformidade.items() if v == 'Não Conforme']
            pdf.set_text_color(220, 38, 38)
            pdf.set_font('Helvetica', 'B', 8)
            if pendencias_mapped:
                pdf.cell(190, 5, "  PENDENCIAS IDENTIFICADAS:", 'LR', 1, 'L')
                pdf.set_font('Helvetica', '', 8)
                for p in pendencias_mapped:
                    pdf.cell(190, 4, f"    - {p}", 'LR', 1, 'L')
            else:
                pdf.cell(190, 5, "  PENDENCIAS IDENTIFICADAS: Sem dados de checklist.", 'LR', 1, 'L')
                
            obs = item.get('observacoes', '').strip()
            if obs:
                pdf.set_text_color(100, 116, 139)
                pdf.set_font('Helvetica', 'I', 8)
                pdf.cell(190, 5, f"  Observacoes: {obs}", 'LR', 1, 'L')
                
            pdf.cell(190, 2, "", 'T', 1, 'L')
            pdf.ln(3)

    # CONFORMES
    if pdf.get_y() + 30 > 275:
        pdf.add_page()
    pdf.ln(3)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_text_color(21, 128, 61)
    pdf.cell(0, 6, f'2. DETALHES DE {titulo.upper()} EM CONFORMIDADE', 0, 1, 'L')
    pdf.ln(2)

    if not conformes:
        pdf.set_font('Helvetica', 'I', 10)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 8, f'Nenhum {titulo.lower()} em conformidade.', 0, 1, 'L')
    else:
        pdf.set_fill_color(220, 252, 231)
        pdf.set_draw_color(134, 239, 172)
        pdf.set_text_color(21, 128, 61)
        pdf.set_font('Helvetica', 'B', 8.5)
        pdf.cell(20, 7, 'Codigo', 1, 0, 'C', True)
        pdf.cell(60, 7, 'Localizacao', 1, 0, 'C', True)
        pdf.cell(35, 7, 'Tipo', 1, 0, 'C', True)
        pdf.cell(25, 7, 'Brigadista', 1, 0, 'C', True)
        pdf.cell(25, 7, 'Vistoria', 1, 0, 'C', True)
        pdf.cell(25, 7, 'Recarga', 1, 1, 'C', True)
        
        pdf.set_text_color(51, 65, 85)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_draw_color(226, 232, 240)
        
        zebra = False
        for item in conformes:
            id_el = item.get('idExtintor', '')
            local = item.get('local', 'N/A')
            tipo = item.get('tipo', 'N/A')
            brigadista = item.get('nomeBrigadista', 'N/A')
            data_insp = item.get('dataInspecao', 'N/A')
            venc_recarga = item.get('vencimentoRecarga', 'N/A')
            try: data_insp = datetime.strptime(data_insp, '%Y-%m-%d').strftime('%d/%m/%Y')
            except: pass
            try: venc_recarga = datetime.strptime(venc_recarga, '%Y-%m-%d').strftime('%d/%m/%Y')
            except: pass
                
            pdf.set_fill_color(248, 250, 252) if zebra else pdf.set_fill_color(255, 255, 255)
            zebra = not zebra
            
            pdf.cell(20, 6, id_el, 1, 0, 'C', True)
            pdf.cell(60, 6, f" {local}", 1, 0, 'L', True)
            pdf.cell(35, 6, f" {tipo}", 1, 0, 'C', True)
            pdf.cell(25, 6, f" {brigadista}", 1, 0, 'L', True)
            pdf.cell(25, 6, data_insp, 1, 0, 'C', True)
            pdf.cell(25, 6, venc_recarga, 1, 1, 'C', True)

print("Gerando relatório PDF...")
pdf = RelatorioPDF()
pdf.alias_nb_pages()
pdf.set_margins(10, 40, 10)

if lista_extintores:
    gerar_secao_pdf(pdf, "Extintores", lista_extintores, MAP_CHECKLIST_EXT)
if lista_hidrantes:
    gerar_secao_pdf(pdf, "Hidrantes", lista_hidrantes, MAP_CHECKLIST_HID)
if not lista_extintores and not lista_hidrantes:
    pdf.add_page()
    pdf.set_font('Helvetica', 'I', 12)
    pdf.cell(0, 10, 'Nenhuma vistoria encontrada.', 0, 1, 'C')

pdf.output("Relatorio_Brigada.pdf")
print("PDF gerado com sucesso: Relatorio_Brigada.pdf")

def formatar_aba_excel(worksheet, df_data):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    row_zebra_fill = PatternFill(start_color='F2F6F9', end_color='F2F6F9', fill_type='solid')
    conf_fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
    nconf_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    for col_idx in range(1, len(df_data.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
    for row_idx in range(2, len(df_data) + 2):
        use_zebra = (row_idx % 2 == 0)
        status_val = worksheet.cell(row=row_idx, column=4).value
        for col_idx in range(1, len(df_data.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = Font(name='Segoe UI', size=10)
            cell.border = thin_border
            if col_idx in [1, 4, 5, 8, 9]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if col_idx == 4:
                if status_val == 'Conforme':
                    cell.fill = conf_fill
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color='375623')
                else:
                    cell.fill = nconf_fill
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color='C65911')
            elif use_zebra:
                cell.fill = row_zebra_fill
                
    worksheet.row_dimensions[1].height = 30
    for r in range(2, len(df_data) + 2):
        worksheet.row_dimensions[r].height = 20
        
    for col in worksheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if col_letter == 'J':
                max_len = max(max_len, min(len(val), 30))
            else:
                max_len = max(max_len, len(val))
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

if lista_extintores or lista_hidrantes:
    print("Gerando planilha Excel estilizada...")
    file_name = "Relatorio_Brigada.xlsx"
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        
        if lista_extintores:
            linhas_ext = []
            for item in lista_extintores:
                is_conf = all(v == "Conforme" for v in item.get('conformidade', {}).values()) if item.get('conformidade') else False
                row_data = {
                    'Código': item.get('idExtintor', ''),
                    'Localização': item.get('local', ''),
                    'Tipo': item.get('tipo', ''),
                    'Status Geral': 'Conforme' if is_conf else 'Não Conforme',
                    'Data da Vistoria': item.get('dataInspecao', ''),
                    'Brigadista': item.get('nomeBrigadista', ''),
                    'E-mail Brigadista': item.get('emailBrigadista', ''),
                    'Vencimento Recarga': item.get('vencimentoRecarga', ''),
                    'Vencimento Teste Hidrostático': item.get('vencimentoHidrostatico', ''),
                    'Observações': item.get('observacoes', '')
                }
                for key_db, label_friendly in MAP_CHECKLIST_EXT.items():
                    row_data[label_friendly] = item.get('conformidade', {}).get(key_db, 'N/A')
                linhas_ext.append(row_data)
            df_ext = pd.DataFrame(linhas_ext)
            df_ext.to_excel(writer, index=False, sheet_name='Extintores')
            formatar_aba_excel(writer.sheets['Extintores'], df_ext)

        if lista_hidrantes:
            linhas_hid = []
            for item in lista_hidrantes:
                is_conf = all(v == "Conforme" for v in item.get('conformidade', {}).values()) if item.get('conformidade') else False
                row_data = {
                    'Código': item.get('idExtintor', ''),
                    'Localização': item.get('local', ''),
                    'Tipo': item.get('tipo', ''),
                    'Status Geral': 'Conforme' if is_conf else 'Não Conforme',
                    'Data da Vistoria': item.get('dataInspecao', ''),
                    'Brigadista': item.get('nomeBrigadista', ''),
                    'E-mail Brigadista': item.get('emailBrigadista', ''),
                    'Vencimento Manutenção': item.get('vencimentoRecarga', ''),
                    'Vencimento Teste Hidrostático': item.get('vencimentoHidrostatico', ''),
                    'Observações': item.get('observacoes', '')
                }
                for key_db, label_friendly in MAP_CHECKLIST_HID.items():
                    row_data[label_friendly] = item.get('conformidade', {}).get(key_db, 'N/A')
                linhas_hid.append(row_data)
            df_hid = pd.DataFrame(linhas_hid)
            df_hid.to_excel(writer, index=False, sheet_name='Hidrantes')
            formatar_aba_excel(writer.sheets['Hidrantes'], df_hid)
            
    print("Planilha Excel gerada com sucesso: Relatorio_Brigada.xlsx")